from __future__ import annotations

from io import BytesIO

import db

BLUE = '315B7D'
LIGHT = 'D9E5F0'
BORDER = 'B8C2CC'


def build_packing_export(case_id: int) -> bytes:
    """엑셀을 실제 생성할 때만 openpyxl을 불러옵니다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def border() -> Border:
        side = Side(style='thin', color=BORDER)
        return Border(top=side, bottom=side, left=side, right=side)

    case = db.row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError('수출 건을 찾을 수 없습니다.')

    items = db.rows(
        '''SELECT s.*, b.length_cm,b.width_cm,b.height_cm,b.weight_kg
           FROM shipment_items s
           LEFT JOIN boxes b ON b.case_id=s.case_id AND b.box_no=s.box_no
           WHERE s.case_id=? AND s.box_no IS NOT NULL
           ORDER BY s.box_no,s.id''',
        (case_id,),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Packing Detail'
    ws.merge_cells('A1:I2')
    ws['A1'] = 'EXPORT PACKING DETAIL'
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    info = [
        ('국가', case['country']),
        ('바이어', case['buyer'] or '-'),
        ('운송방식', case['transport_mode']),
        ('국내배송', case['domestic_method'] or '-'),
    ]
    for row_no, (label, value) in enumerate(info, 4):
        ws[f'A{row_no}'] = label
        ws[f'B{row_no}'] = value
        ws.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=5)
        ws[f'A{row_no}'].fill = PatternFill('solid', fgColor=LIGHT)
        ws[f'A{row_no}'].font = Font(bold=True)

    detail = ''
    if case['domestic_method'] == '로젠택배':
        detail = f"송장번호: {case['tracking_no']}"
    elif case['domestic_method'] == '퀵배송':
        detail = f"기사: {case['driver_name']} / {case['driver_phone']}"

    ws['F7'] = '배송 상세'
    ws['G7'] = detail
    ws.merge_cells('G7:I7')
    ws['F7'].fill = PatternFill('solid', fgColor=LIGHT)
    ws['F7'].font = Font(bold=True)

    headers = ['박스번호', '사업장', '로케이션', '제품명', 'LOT', '유통기한', '수량', '무게(kg)', '박스 사이즈(cm)']
    header_row = 10
    for column, header in enumerate(headers, 1):
        cell = ws.cell(header_row, column, header)
        cell.fill = PatternFill('solid', fgColor=BLUE)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border()

    groups: dict[int, list] = {}
    for item in items:
        groups.setdefault(item['box_no'], []).append(item)

    current_row = header_row + 1
    for box_no, group in groups.items():
        first_row = current_row
        for item in group:
            box_size = f"{item['length_cm']:g} x {item['width_cm']:g} x {item['height_cm']:g}"
            values = [
                box_no,
                item['business_unit'],
                item['location'],
                item['product_name'],
                item['lot_no'],
                item['expiry_date'],
                item['requested_qty'],
                item['weight_kg'],
                box_size,
            ]
            for column, value in enumerate(values, 1):
                cell = ws.cell(current_row, column, value)
                cell.border = border()
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            current_row += 1

        last_row = current_row - 1
        if last_row > first_row:
            for column in (1, 8, 9):
                ws.merge_cells(start_row=first_row, start_column=column, end_row=last_row, end_column=column)
        for column in (1, 8, 9):
            ws.cell(first_row, column).alignment = Alignment(horizontal='center', vertical='center')

    widths = [11, 14, 14, 32, 16, 14, 10, 12, 22]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.freeze_panes = 'A11'
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
