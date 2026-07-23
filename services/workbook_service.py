from __future__ import annotations

from pathlib import Path

import db


SECTION_TITLES = {'기본 정보', '주문 목록', '출고 진행 상황', '국내배송 정보'}


def _style_sheet(ws, widths: dict[str, float], header_rows: set[int] | None = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_rows = header_rows or set()
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    section_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='B8C2CC')
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx in range(1, ws.max_row + 1):
        first = ws.cell(row_idx, 1)
        if first.value in SECTION_TITLES:
            for cell in ws[row_idx]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
        elif row_idx in header_rows:
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = 'A2'


def _box_size_text(row) -> str:
    values = [row['length_cm'], row['width_cm'], row['height_cm']]
    if not all(value not in (None, '') for value in values):
        return ''
    return ' × '.join(f'{float(value):g}' for value in values)


def write_case_workbook(case_id: int, folder: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    case = db.row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError(f'수출 건을 찾을 수 없습니다: {case_id}')
    folder.mkdir(parents=True, exist_ok=True)
    workbook_path = folder / '수출진행내역.xlsx'
    orders = db.rows(
        'SELECT id, product_name, quantity, unit, created_at FROM order_items WHERE case_id=? ORDER BY id',
        (case_id,),
    )
    shipments = db.rows(
        '''SELECT o.id AS order_item_id, o.product_name AS order_product_name,
                  o.quantity AS order_quantity, o.unit,
                  s.id AS shipment_id, s.business_unit,
                  s.product_name AS actual_product_name, s.lot_no,
                  s.expiry_date, s.requested_qty, s.box_no, s.updated_at,
                  b.length_cm, b.width_cm, b.height_cm, b.weight_kg
           FROM order_items o
           LEFT JOIN shipment_items s ON s.order_item_id=o.id
           LEFT JOIN boxes b ON b.case_id=s.case_id AND b.box_no=s.box_no
           WHERE o.case_id=?
           ORDER BY o.id,
                    CASE WHEN s.box_no IS NULL THEN 1 ELSE 0 END,
                    s.box_no,
                    s.id''',
        (case_id,),
    )

    wb = Workbook()
    ws1 = wb.active
    ws1.title = '주문 접수 내역'
    ws1.append(['주문 접수 내역'])
    ws1.append(['기본 정보'])
    ws1.append(['수출번호', '국가', '바이어', '운송방식', '진행단계', '상태', '비고', '생성일', '수정일'])
    ws1.append([case['export_no'], case['country'], case['buyer'], case['transport_mode'], case['stage'], case['status'], case['note'], case['created_at'], case['updated_at']])
    ws1.append([])
    ws1.append(['주문 목록'])
    ws1.append(['제품명', '수량', '단위', '등록일'])
    for item in orders:
        ws1.append([item['product_name'], item['quantity'], item['unit'], item['created_at']])
    _style_sheet(
        ws1,
        {'A': 28, 'B': 14, 'C': 18, 'D': 14, 'E': 16, 'F': 12, 'G': 30, 'H': 20, 'I': 20},
        {3, 7},
    )

    ws2 = wb.create_sheet('출고 진행 상황')
    ws2.append(['출고 진행 상황'])
    ws2.append([
        '주문제품', '총 주문수량', '단위', '사업장', '실출고 제품명',
        '제조번호', '유통기한', '출고수량', '박스번호',
        '박스 사이즈 (가로 × 세로 × 높이, cm)', 'GW (kg)', '수정일',
    ])

    grouped: dict[int, list] = {int(order['id']): [] for order in orders}
    for shipment in shipments:
        if shipment['shipment_id'] is not None:
            grouped.setdefault(int(shipment['order_item_id']), []).append(shipment)

    box_rows: dict[int, list[int]] = {}
    order_fill = PatternFill('solid', fgColor='F5F8FB')

    for order in orders:
        order_row = ws2.max_row + 1
        ws2.append([order['product_name'], order['quantity'], order['unit'], '', '', '', '', '', '', '', '', ''])
        for cell in ws2[order_row]:
            cell.font = Font(bold=True)
            cell.fill = order_fill

        order_shipments = grouped.get(int(order['id']), [])
        for shipment in order_shipments:
            box_no = int(shipment['box_no']) if shipment['box_no'] is not None else None
            row_no = ws2.max_row + 1
            ws2.append([
                f'└ {shipment["actual_product_name"] or order["product_name"]}',
                '',
                order['unit'],
                shipment['business_unit'] or '',
                shipment['actual_product_name'] or '',
                shipment['lot_no'] or '',
                shipment['expiry_date'] or '',
                shipment['requested_qty'] or 0,
                box_no or '',
                _box_size_text(shipment),
                shipment['weight_kg'] if shipment['weight_kg'] not in (None, '') else '',
                shipment['updated_at'] or '',
            ])
            ws2.cell(row_no, 1).alignment = Alignment(indent=1, vertical='center', wrap_text=True)
            if box_no is not None:
                box_rows.setdefault(box_no, []).append(row_no)

    for box_no in sorted(box_rows):
        rows = box_rows[box_no]
        if len(rows) > 1:
            for column in (9, 10, 11):
                ws2.merge_cells(start_row=rows[0], start_column=column, end_row=rows[-1], end_column=column)
        for column in (9, 10, 11):
            ws2.cell(rows[0], column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    _style_sheet(
        ws2,
        {
            'A': 32, 'B': 15, 'C': 10, 'D': 16, 'E': 28, 'F': 18,
            'G': 16, 'H': 14, 'I': 12, 'J': 28, 'K': 12, 'L': 20,
        },
        {2},
    )

    ws3 = wb.create_sheet('국내배송 정보')
    ws3.append(['국내배송 정보'])
    ws3.append(['항목', '내용'])
    for label, value in [
        ('국내배송 방식', case['domestic_method']), ('국내배송 일자', case['actual_ship_date']),
        ('송장번호', case['tracking_no']), ('배송기사 이름', case['driver_name']),
        ('배송기사 연락처', case['driver_phone']), ('현재 단계', case['stage']),
        ('상태', case['status']), ('비고', case['note']), ('취소 사유', case['cancel_reason']),
        ('취소 일시', case['cancelled_at']), ('최종 수정일', case['updated_at']),
    ]:
        ws3.append([label, value or ''])
    _style_sheet(ws3, {'A': 24, 'B': 48}, {2})
    wb.save(workbook_path)
    return workbook_path
