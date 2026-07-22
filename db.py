from __future__ import annotations

import shutil
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / 'export.db'
UPLOAD_DIR = BASE_DIR / 'uploads'
STAGES = ['주문 접수','제품 준비','실출고 입력','패킹','국내배송','선적 준비','선적 완료','완료','취소']
TRANSPORT_MODES = ['AIR','SEA','HAND']
INVALID_FOLDER_CHARS = '\\/:*?"<>|'


@contextmanager
def connect() -> Iterable[sqlite3.Connection]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def now_text() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {r['name'] for r in conn.execute(f'PRAGMA table_info({table})')}


def _add_column(conn: sqlite3.Connection, table: str, definition: str) -> None:
    name = definition.split()[0]
    if name not in _columns(conn, table):
        conn.execute(f'ALTER TABLE {table} ADD COLUMN {definition}')


def _remove_expected_ship_date_column(conn: sqlite3.Connection) -> None:
    columns = _columns(conn, 'export_cases')
    if 'expected_ship_date' not in columns:
        return
    conn.executescript('''
    PRAGMA foreign_keys = OFF;
    CREATE TABLE export_cases_new (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        export_no TEXT NOT NULL UNIQUE,
        buyer TEXT DEFAULT '',
        country TEXT NOT NULL DEFAULT '',
        transport_mode TEXT DEFAULT 'AIR',
        stage TEXT NOT NULL DEFAULT '주문 접수',
        status TEXT NOT NULL DEFAULT '진행중',
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        domestic_method TEXT DEFAULT '',
        tracking_no TEXT DEFAULT '',
        driver_name TEXT DEFAULT '',
        driver_phone TEXT DEFAULT '',
        note TEXT DEFAULT '',
        actual_ship_date TEXT DEFAULT '',
        folder_path TEXT DEFAULT '',
        cancel_reason TEXT DEFAULT '',
        cancelled_at TEXT DEFAULT '',
        previous_stage TEXT DEFAULT ''
    );
    INSERT INTO export_cases_new(
        id,export_no,buyer,country,transport_mode,stage,status,created_at,updated_at,
        domestic_method,tracking_no,driver_name,driver_phone,note,actual_ship_date,
        folder_path,cancel_reason,cancelled_at,previous_stage
    )
    SELECT
        id,export_no,buyer,country,transport_mode,stage,status,created_at,updated_at,
        COALESCE(domestic_method,''),COALESCE(tracking_no,''),COALESCE(driver_name,''),
        COALESCE(driver_phone,''),COALESCE(note,''),COALESCE(actual_ship_date,''),
        COALESCE(folder_path,''),COALESCE(cancel_reason,''),COALESCE(cancelled_at,''),
        COALESCE(previous_stage,'')
    FROM export_cases;
    DROP TABLE export_cases;
    ALTER TABLE export_cases_new RENAME TO export_cases;
    PRAGMA foreign_keys = ON;
    ''')


def init_db() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript('''
        CREATE TABLE IF NOT EXISTS export_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            export_no TEXT NOT NULL UNIQUE,
            buyer TEXT DEFAULT '',
            country TEXT NOT NULL DEFAULT '',
            transport_mode TEXT DEFAULT 'AIR',
            stage TEXT NOT NULL DEFAULT '주문 접수',
            status TEXT NOT NULL DEFAULT '진행중',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL REFERENCES export_cases(id) ON DELETE CASCADE,
            product_name TEXT NOT NULL,
            quantity REAL NOT NULL DEFAULT 0,
            unit TEXT DEFAULT 'EA',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS shipment_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL REFERENCES export_cases(id) ON DELETE CASCADE,
            business_unit TEXT DEFAULT '',
            location TEXT DEFAULT '',
            product_name TEXT NOT NULL,
            lot_no TEXT DEFAULT '',
            expiry_date TEXT DEFAULT '',
            requested_qty REAL NOT NULL DEFAULT 0,
            box_no INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS boxes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL REFERENCES export_cases(id) ON DELETE CASCADE,
            box_no INTEGER NOT NULL,
            length_cm REAL DEFAULT 0,
            width_cm REAL DEFAULT 0,
            height_cm REAL DEFAULT 0,
            weight_kg REAL DEFAULT 0,
            updated_at TEXT NOT NULL,
            UNIQUE(case_id, box_no)
        );
        CREATE TABLE IF NOT EXISTS attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL REFERENCES export_cases(id) ON DELETE CASCADE,
            file_name TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            category TEXT DEFAULT '출고사진',
            uploaded_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER REFERENCES export_cases(id) ON DELETE CASCADE,
            action TEXT NOT NULL,
            detail TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL
        );
        ''')
        for definition in [
            "domestic_method TEXT DEFAULT ''",
            "tracking_no TEXT DEFAULT ''",
            "driver_name TEXT DEFAULT ''",
            "driver_phone TEXT DEFAULT ''",
            "note TEXT DEFAULT ''",
            "actual_ship_date TEXT DEFAULT ''",
            "folder_path TEXT DEFAULT ''",
            "cancel_reason TEXT DEFAULT ''",
            "cancelled_at TEXT DEFAULT ''",
            "previous_stage TEXT DEFAULT ''",
        ]:
            _add_column(conn, 'export_cases', definition)
        _remove_expected_ship_date_column(conn)


def rows(query: str, params: tuple[Any, ...] = ()) -> list[sqlite3.Row]:
    with connect() as conn:
        return list(conn.execute(query, params).fetchall())


def row(query: str, params: tuple[Any, ...] = ()) -> sqlite3.Row | None:
    with connect() as conn:
        return conn.execute(query, params).fetchone()


def execute(query: str, params: tuple[Any, ...] = ()) -> int:
    with connect() as conn:
        cur = conn.execute(query, params)
        return int(cur.lastrowid or 0)


def executemany(query: str, values: list[tuple[Any, ...]]) -> None:
    with connect() as conn:
        conn.executemany(query, values)

    if 'insert into order_items' in ' '.join(query.lower().split()):
        case_ids = {int(value[0]) for value in values if value}
        for case_id in case_ids:
            sync_case_folder(case_id)


def get_setting(key: str, default: str = '') -> str:
    result = row('SELECT value FROM settings WHERE key=?', (key,))
    return str(result['value']) if result else default


def set_setting(key: str, value: str) -> None:
    execute(
        '''INSERT INTO settings(key,value,updated_at) VALUES (?,?,?)
           ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at''',
        (key, value, now_text()),
    )


def storage_root() -> Path:
    configured = get_setting('shared_root').strip()
    return Path(configured).expanduser() if configured else UPLOAD_DIR


def test_storage_root(path_text: str) -> tuple[bool, str]:
    path_text = path_text.strip()
    if not path_text:
        return False, '폴더 경로를 입력하세요.'
    try:
        path = Path(path_text).expanduser()
        path.mkdir(parents=True, exist_ok=True)
        probe = path / '.export_write_test'
        probe.write_text('ok', encoding='utf-8')
        probe.unlink()
        return True, str(path.resolve())
    except Exception as exc:
        return False, str(exc)


def add_history(case_id: int | None, action: str, detail: str) -> None:
    execute('INSERT INTO history(case_id, action, detail, created_at) VALUES (?,?,?,?)',
            (case_id, action, detail, now_text()))


def next_export_no() -> str:
    year = datetime.now().year
    result = row('SELECT export_no FROM export_cases WHERE export_no LIKE ? ORDER BY export_no DESC LIMIT 1',
                 (f'EXP-{year}-%',))
    if not result:
        return f'EXP-{year}-001'
    try:
        number = int(result['export_no'].split('-')[-1]) + 1
    except ValueError:
        number = 1
    return f'EXP-{year}-{number:03d}'


def active_cases(country: str | None = None) -> list[sqlite3.Row]:
    sql = "SELECT * FROM export_cases WHERE status='진행중' AND stage NOT IN ('완료','취소')"
    params: tuple[Any, ...] = ()
    if country:
        sql += ' AND country=?'
        params = (country,)
    return rows(sql + ' ORDER BY created_at', params)


def parse_date(value: str | None) -> datetime | None:
    if not value:
        return None
    for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def sanitize_folder_part(value: str | None, fallback: str = '미입력') -> str:
    text = str(value or '').strip() or fallback
    for char in INVALID_FOLDER_CHARS:
        text = text.replace(char, '_')
    text = ' '.join(text.split())
    return text.strip(' .') or fallback


def order_item_summary(case_id: int) -> str:
    product_rows = rows(
        'SELECT product_name FROM order_items WHERE case_id=? ORDER BY id',
        (case_id,),
    )
    products: list[str] = []
    seen: set[str] = set()
    for product in product_rows:
        name = sanitize_folder_part(product['product_name'], '')
        if name and name not in seen:
            products.append(name)
            seen.add(name)

    if not products:
        return ''
    if len(products) <= 2:
        return ', '.join(products)
    return f'{products[0]}, {products[1]} 외 {len(products) - 2}품목'


def case_folder_name(case: sqlite3.Row | dict[str, Any]) -> str:
    case_id = int(case['id'])
    country = sanitize_folder_part(case['country'], '국가미입력')
    buyer = sanitize_folder_part(case['buyer'], '')
    transport = sanitize_folder_part(case['transport_mode'], '')
    summary = order_item_summary(case_id)

    parts = [country]
    if buyer:
        parts.append(buyer)
    if transport:
        parts.append(transport)
    if summary:
        parts.append(summary)

    name = '_'.join(parts)
    actual_ship_date = parse_date(
        case['actual_ship_date'] if 'actual_ship_date' in case.keys() else ''
    )
    domestic_method = str(
        case['domestic_method'] if 'domestic_method' in case.keys() else ''
    ).strip()
    if actual_ship_date and domestic_method:
        name = f'{actual_ship_date.strftime("%m%d")}_{name}'

    status = str(case['status'] if 'status' in case.keys() else '')
    stage = str(case['stage'] if 'stage' in case.keys() else '')
    if status == '취소' or stage == '취소':
        return name if name.startswith('[취소]') else f'[취소]{name}'
    return name.removeprefix('[취소]')


def case_folder_base(case: sqlite3.Row | dict[str, Any]) -> Path:
    actual = parse_date(case['actual_ship_date'] if 'actual_ship_date' in case.keys() else '')
    year = (actual or datetime.now()).strftime('%Y')
    country = sanitize_folder_part(case['country'], '국가미입력')
    return storage_root() / country / year


def unique_folder_path(base: Path, folder_name: str, current_path: Path | None = None) -> Path:
    target = base / folder_name
    if current_path and target.resolve() == current_path.resolve():
        return target
    if not target.exists():
        return target
    idx = 2
    while True:
        candidate = base / f'{folder_name}_{idx}'
        if current_path and candidate.resolve() == current_path.resolve():
            return candidate
        if not candidate.exists():
            return candidate
        idx += 1


def _style_sheet(ws, widths: dict[str, float]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    section_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='B8C2CC')

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in range(1, ws.max_row + 1):
        first = ws.cell(row_idx, 1)
        if first.value in {'기본 정보', '주문 목록', '실출고 진행 상황', '국내배송 정보'}:
            for cell in ws[row_idx]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
        elif row_idx in {3, 11}:
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = Font(bold=True)

    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = 'A2'


def write_case_workbook(case_id: int, folder: Path | None = None) -> Path:
    from openpyxl import Workbook

    case = row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError(f'수출 건을 찾을 수 없습니다: {case_id}')

    target_folder = folder or ensure_case_folder(case_id)
    target_folder.mkdir(parents=True, exist_ok=True)
    workbook_path = target_folder / '수출진행내역.xlsx'

    orders = rows(
        'SELECT product_name, quantity, unit, created_at FROM order_items WHERE case_id=? ORDER BY id',
        (case_id,),
    )
    shipments = rows(
        '''SELECT o.product_name AS order_product_name, o.quantity AS order_quantity, o.unit,
                  s.business_unit, s.product_name AS actual_product_name, s.lot_no,
                  s.expiry_date, s.requested_qty, s.box_no, s.updated_at
           FROM order_items o
           LEFT JOIN shipment_items s ON s.order_item_id=o.id
           WHERE o.case_id=?
           ORDER BY o.id, s.id''',
        (case_id,),
    )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = '주문 접수 내역'
    ws1.append(['주문 접수 내역'])
    ws1.append(['기본 정보'])
    ws1.append(['수출번호', '국가', '바이어', '운송방식', '진행단계', '상태', '비고', '생성일', '수정일'])
    ws1.append([
        case['export_no'], case['country'], case['buyer'], case['transport_mode'],
        case['stage'], case['status'], case['note'], case['created_at'], case['updated_at'],
    ])
    ws1.append([])
    ws1.append(['주문 목록'])
    ws1.append(['제품명', '수량', '단위', '등록일'])
    for item in orders:
        ws1.append([item['product_name'], item['quantity'], item['unit'], item['created_at']])
    _style_sheet(ws1, {'A': 28, 'B': 14, 'C': 18, 'D': 14, 'E': 16, 'F': 12, 'G': 30, 'H': 20, 'I': 20})

    ws2 = wb.create_sheet('실출고 진행 상황')
    ws2.append(['실출고 진행 상황'])
    ws2.append(['주문제품', '주문수량', '단위', '사업장', '실제 제품명', '제조번호', '유통기한', '출고수량', '박스번호', '수정일'])
    for item in shipments:
        ws2.append([
            item['order_product_name'], item['order_quantity'], item['unit'],
            item['business_unit'] or '', item['actual_product_name'] or '', item['lot_no'] or '',
            item['expiry_date'] or '', item['requested_qty'] or 0, item['box_no'] or '', item['updated_at'] or '',
        ])
    _style_sheet(ws2, {'A': 28, 'B': 14, 'C': 10, 'D': 16, 'E': 28, 'F': 18, 'G': 16, 'H': 14, 'I': 12, 'J': 20})

    ws3 = wb.create_sheet('국내배송 정보')
    ws3.append(['국내배송 정보'])
    ws3.append(['항목', '내용'])
    delivery_rows = [
        ('국내배송 방식', case['domestic_method']),
        ('국내배송 일자', case['actual_ship_date']),
        ('송장번호', case['tracking_no']),
        ('배송기사 이름', case['driver_name']),
        ('배송기사 연락처', case['driver_phone']),
        ('현재 단계', case['stage']),
        ('상태', case['status']),
        ('비고', case['note']),
        ('취소 사유', case['cancel_reason']),
        ('취소 일시', case['cancelled_at']),
        ('최종 수정일', case['updated_at']),
    ]
    for label, value in delivery_rows:
        ws3.append([label, value or ''])
    _style_sheet(ws3, {'A': 24, 'B': 48})

    wb.save(workbook_path)
    return workbook_path


def ensure_case_folder(case_id: int) -> Path:
    case = row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError(f'수출 건을 찾을 수 없습니다: {case_id}')
    saved_path = Path(case['folder_path']) if case['folder_path'] else None
    if saved_path and saved_path.exists():
        write_case_workbook(case_id, saved_path)
        return saved_path
    base = case_folder_base(case)
    base.mkdir(parents=True, exist_ok=True)
    target = unique_folder_path(base, case_folder_name(case))
    target.mkdir(parents=True, exist_ok=True)
    execute('UPDATE export_cases SET folder_path=?,updated_at=? WHERE id=?', (str(target), now_text(), case_id))
    write_case_workbook(case_id, target)
    return target


def refresh_attachment_paths(case_id: int, old_root: Path, new_root: Path) -> None:
    old_text = str(old_root)
    for attachment in rows('SELECT id, stored_path FROM attachments WHERE case_id=?', (case_id,)):
        stored = str(attachment['stored_path'])
        if stored.startswith(old_text):
            replacement = str(new_root / Path(stored).relative_to(old_root))
            execute('UPDATE attachments SET stored_path=? WHERE id=?', (replacement, attachment['id']))


def sync_case_folder(case_id: int) -> Path:
    case = row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError(f'수출 건을 찾을 수 없습니다: {case_id}')
    base = case_folder_base(case)
    base.mkdir(parents=True, exist_ok=True)
    current = Path(case['folder_path']) if case['folder_path'] else None
    target = unique_folder_path(base, case_folder_name(case), current)

    if current and current.exists():
        if current.resolve() != target.resolve():
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(current), str(target))
            refresh_attachment_paths(case_id, current, target)
    else:
        target.mkdir(parents=True, exist_ok=True)

    execute('UPDATE export_cases SET folder_path=?,updated_at=? WHERE id=?', (str(target), now_text(), case_id))
    write_case_workbook(case_id, target)
    return target


def rebuild_all_case_folders() -> list[tuple[int, Path]]:
    results: list[tuple[int, Path]] = []
    for case in rows('SELECT id FROM export_cases ORDER BY id'):
        results.append((int(case['id']), sync_case_folder(int(case['id']))))
    return results


def move_file_to_case(case_id: int, source: Path, category: str = '출고사진') -> Path:
    folder = ensure_case_folder(case_id)
    source = Path(source)
    destination = folder / source.name
    counter = 2
    while destination.exists():
        destination = folder / f'{source.stem}_{counter}{source.suffix}'
        counter += 1
    shutil.move(str(source), str(destination))
    execute(
        'INSERT INTO attachments(case_id,file_name,stored_path,category,uploaded_at) VALUES (?,?,?,?,?)',
        (case_id, destination.name, str(destination), category, now_text()),
    )
    return destination
